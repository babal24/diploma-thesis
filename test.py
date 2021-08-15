#Implementation of file analysis 
#Install:
# pip install scipy++++++
# pip install numpy
# pip install pandas
# pip install PySimpleGUI==3.20.0
# pip install -U scikit-learn
# pip install numpy


import PySimpleGUI as sg      
import sys   
import pandas as pd   
from sklearn import linear_model
import numpy as np
from array import array
from openpyxl import workbook
from openpyxl import load_workbook
import os.path
import ntpath
from math import pi
from openpyxl.chart import (
    RadarChart,
    Reference,
    Series
)



ntpath.basename("a/b/c")

def path_leaf(path):
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)


if len(sys.argv) == 1:      
    event, ( fname,) = sg.Window('My Script').Layout([[sg.Text('Choose files to analyze:')],      
                                                        [sg.In(), sg.FilesBrowse()],      
                                                        [sg.CloseButton('Execute Multivariate Regression'), sg.CloseButton('Cancel')], 
                                                        [sg.Quit(button_color=('black', 'orange'))]]).Read()  
    if event == 'Quit':
        sg.Popup("Bye!") 
        raise SystemExit("User quit")  
                                                             
else:      
        fname = sys.argv[1]      
      
if not fname:      
    sg.Popup("Cancel", "No filename supplied")      
    raise SystemExit("Cancelling: no filename supplied")      

  
inputs = fname.split(';')
#print (inputs)
i=1
prevStudentAEM = 'first'

for inputFileDir in inputs:
    xlsFileInput = path_leaf(inputFileDir)
    #print(xlsFileInput)
    extensionRemovedName = xlsFileInput.split(".")[0]
    StudentAEMandRun = extensionRemovedName.split("_")
    StudentAEM = StudentAEMandRun[0]
    Run = StudentAEMandRun[1]
    #print(StudentAEM)
    #print(Run)
    curr_dataframe = pd.read_excel(inputFileDir, index=0) 
    X = curr_dataframe[['απόθεμα μπύρας','κόστος διατήρησης αποθέματος μπύρας','ολοκλήρωση ωρίμανσης μπύρας','χρόνος παραγωγής του προιόντος',
    'εβδομαδιαίες χαμένες πωλήσεις','εβδομαδιαία έσοδα','ισοζύγιο', 'εβδομαδιαίες πωλήσεις','τιμή βαρελιού μπύρας',
    'εβδομαδιαίο κόστος','συνολικές χαμένες πωλήσεις', 'κόστος διατήρησης αποθέματος Α υλών', 'παραλαβή Α υλών', 'χρόνος παράδοσης Α υλών', 'κόστος αγοράς Α υλών', 
    'κόστος παραγωγής', 'απόθεμα Α υλών', 'σταθερο κόστος λειτουργίας', 'εβδομαδιαία ζήτηση', 'περιθώριο κέρδους', 'συνολικές απώλειες εσόδων', 
    'παραγωγική διαδικασία', 'παραγωγική ικανότητα προμηθευτή', 'εβδομαδιαίες απώλειες απο τις χαμένες πωλήσες' ]] 
    Y = curr_dataframe[['απόφαση για παραγωγή' ,'απόφαση για παραγγελία Α υλών']]
    
    # with sklearn
    regr = linear_model.LinearRegression()
    regr.fit(X, Y)
    np.set_printoptions(suppress=True)


    #print('Intercept: \n', regr.intercept_.round(decimals=8))
    #print('Coefficients: \n', regr.coef_.round(decimals=8))

    #print (type(regr.coef_))
    ## convert your array into a dataframe
    df = pd.DataFrame(regr.coef_.round(decimals=8), index=['απόφαση για παραγωγή', 'απόφαση για παραγγελία Α υλών'], columns=['απόθεμα μπύρας','κόστος διατήρησης αποθέματος μπύρας','ολοκλήρωση ωρίμανσης μπύρας','χρόνος παραγωγής του προιόντος',
    'εβδομαδιαίες χαμένες πωλήσεις','εβδομαδιαία έσοδα','ισοζύγιο', 'εβδομαδιαίες πωλήσεις','τιμή βαρελιού μπύρας',
    'εβδομαδιαίο κόστος','συνολικές χαμένες πωλήσεις', 'κόστος διατήρησης αποθέματος Α υλών', 'παραλαβή Α υλών', 'χρόνος παράδοσης Α υλών', 'κόστος αγοράς Α υλών', 
    'κόστος παραγωγής', 'απόθεμα Α υλών', 'σταθερο κόστος λειτουργίας', 'εβδομαδιαία ζήτηση', 'περιθώριο κέρδους', 'συνολικές απώλειες εσόδων', 
    'παραγωγική διαδικασία', 'παραγωγική ικανότητα προμηθευτή', 'εβδομαδιαίες απώλειες απο τις χαμένες πωλήσες' ])


    # filepath = 'output.xlsx'
    # if filename:
    if StudentAEM != prevStudentAEM :
        i = 1

    filepath = StudentAEM + '.xlsx'
    if i == 1 :
        ## save to xlsx file
        sheetid = "run_" + Run
        df.to_excel(filepath, sheet_name=sheetid, index=True)    
        sg.Popup("File created", "You can access it in script directory")
    else:
        
        sheetid = "run_" + Run
        book = load_workbook(filepath)
        writer = pd.ExcelWriter(filepath, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, sheet_name=sheetid)
        writer.save()

    
    book = load_workbook(filepath)
    curr_ws = book[sheetid]
    chart = RadarChart(radarStyle='standard')
    #labels = Reference(curr_ws, min_col=1, min_row=1, max_row=2)
    data = Reference(curr_ws, min_row=2, max_row=2, min_col=2, max_col=25)
    print(data)
    series = Series(data, title='απόφαση για παραγωγή')
    # chart.add_data(data, titles_from_data=True)
    chart.append(series)
    data2 = Reference(curr_ws, min_row=3, max_row=3, min_col=2, max_col=25)
    series = Series(data2, title="απόφαση για παραγγελία Α υλών")
    chart.append(series)
    # chart.set_categories(labels)
    chart.style = 26
    chart.title = "Decision analysis chart"
    chart.y_axis.delete = True
    #set x-axis
    labels = Reference(curr_ws, min_row=1, max_row=1, min_col=2, max_col=25)
    chart.set_categories(labels)

    curr_ws.add_chart(chart, "B6")
    book.save(filepath)


    prevStudentAEM = StudentAEM
    i += 1

